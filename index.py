#!/usr/bin/env python
# -*- coding: utf-8 -*-

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
import os
from os import sys

def get_all_files(in_dir):
    file_pathes = []
    for root, dirs, files in os.walk(in_dir):
        for file_name in files:
            file_pathes.append(''.join([in_dir, '/', file_name]))
    return file_pathes

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def csv_from_excel(excel_file, sheets, out_dir):
    excel_file_name = excel_file.rsplit('/', 1)[1].split('.')[0];
    print("File " + excel_file_name + " ...")
    print("")
    workbook = load_workbook(excel_file,data_only=True)
    for worksheet_name in sheets:
        print("  Sheet " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)

        dir_path = ''.join([out_dir.decode('utf_8'), '/', excel_file_name.decode('utf_8')]);
        if not os.path.exists(dir_path):
            os.makedirs(dir_path);
        file_path = ''.join([dir_path, '/', worksheet_name,'.csv']);
        print("  Save as " + file_path + " ...")
        your_csv_file = open(file_path, 'wb')
        wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                if isinstance(cell.value, unicode):
                    lrow.append(cell.value.encode('utf_8'))
                else:
                    lrow.append(cell.value)
            wr.writerow(lrow)
        print("  ... done")
        print("")
    workbook.close()

if len(sys.argv) != 3:
	print("Call with " + sys.argv[0] + " <in dir> <out dir>")
	sys.exit(1)
else:
    in_dir = sys.argv[1]
    out_dir = sys.argv[2]
    files = get_all_files(in_dir)
    print files
    for file_path in files:
        sheets = get_all_sheets(file_path)
        assert(sheets != None and len(sheets) > 0)
        csv_from_excel(file_path, sheets, out_dir)
